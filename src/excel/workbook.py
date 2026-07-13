from openpyxl import Workbook

from openpyxl.utils import get_column_letter

from openpyxl.worksheet.table import (
    Table,
    TableStyleInfo
)

from .styles import (
    HEADER_FILL,
    HEADER_FONT,
    THIN_BORDER,
    CENTER
)

# --------------------------------------------------
# Workbook
# --------------------------------------------------

def create_workbook():

    wb = Workbook()

    default = wb.active

    wb.remove(default)

    wb.create_sheet("Home")

    wb.create_sheet("Query")

    wb.create_sheet("Compliance Database")

    wb.create_sheet("Acts")

    wb.create_sheet("Regulators")

    wb.create_sheet("Topics")

    lists = wb.create_sheet("Lists")
    lists.sheet_state = "hidden"

    # Hidden calculation sheet backing the Query engine (match flag +
    # running rank per section).  Keeps the query working in every Excel
    # version — no dynamic-array functions.
    engine = wb.create_sheet("Engine")
    engine.sheet_state = "hidden"

    # Recalculate on open so the query reflects the current selections.
    wb.calculation.fullCalcOnLoad = True

    return wb

# --------------------------------------------------
# Worksheet Styling
# --------------------------------------------------

def style_header(ws):

    for cell in ws[1]:

        cell.fill = HEADER_FILL

        cell.font = HEADER_FONT

        cell.border = THIN_BORDER

        cell.alignment = CENTER

# --------------------------------------------------
# Auto Width
# --------------------------------------------------

def autofit_columns(

    ws,

    minimum=12,

    maximum=60

):

    for column_cells in ws.columns:

        column = get_column_letter(

            column_cells[0].column

        )

        length = 0

        for cell in column_cells:

            try:

                if cell.value:

                    length = max(

                        length,

                        len(str(cell.value))

                    )

            except:

                pass

        ws.column_dimensions[column].width = min(

            max(

                length + 2,

                minimum

            ),

            maximum

        )

# --------------------------------------------------
# Excel Table
# --------------------------------------------------

def make_table(

    ws,

    table_name

):

    end_col = get_column_letter(

        ws.max_column

    )

    end_row = ws.max_row

    table = Table(

        displayName=table_name,

        ref=f"A1:{end_col}{end_row}"

    )

    style = TableStyleInfo(

        name="TableStyleMedium2",

        showFirstColumn=False,

        showLastColumn=False,

        showRowStripes=True,

        showColumnStripes=False

    )

    table.tableStyleInfo = style

    ws.add_table(table)

    ws.freeze_panes = "A2"

# --------------------------------------------------
# Apply common formatting
# --------------------------------------------------

def finish_sheet(

    ws,

    table_name

):

    style_header(ws)

    make_table(ws, table_name)

    autofit_columns(ws)