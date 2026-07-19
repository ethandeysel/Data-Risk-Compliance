"""
Builds the DTIA workbook from a loaded DataLoader.

Sheets
------
Home                 — overview + instructions
Query                — interactive DTIA query engine (dropdowns + FILTER)
Compliance Database  — one row per section, with page references
Acts / Regulators / Topics — rollups
Lists (hidden)       — dropdown source values
"""

import math
import os

from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.hyperlink import Hyperlink

# Excel's hard ceiling on row height (points).
_EXCEL_MAX_ROW_HEIGHT = 409

from .workbook import finish_sheet
from .styles import (
    TITLE_FONT, TITLE_FILL, HEADER_FILL, HEADER_FONT, WRAP, THIN_BORDER,
)
from .loaders import SOURCE_LINK, _pdf_url

# Link text for the Source column — matches where the link points.
_SOURCE_LABEL = "Open PDF" if SOURCE_LINK == "pdf" else "Go to Act"

WRAP_COLUMNS = {
    "Heading", "Topics", "Data Types", "Summary", "DTIA Summary",
    "Requirements", "Requirements (2)", "Source Quote",
}


def _col_width(name):
    """Display width per column — the two Requirements columns are wide
    enough that each one-line requirement fits without wrapping."""
    if name in ("Requirements", "Requirements (2)"):
        return 78
    if name in WRAP_COLUMNS:
        return 45
    return 18


def _result_formula(header, match, col, last):
    """The formula for one result cell, pulled from the Compliance Database
    by the given MATCH.  The Source column becomes a clickable PDF link;
    everything else is a plain INDEX (IF/ISNA/INDEX/MATCH only — no
    functions newer than 2003)."""
    index = f"INDEX('Compliance Database'!${col}$2:${col}${last},{match})"
    if header == "Source":
        # acts mode: the cell holds the Acts-sheet row number, so build a
        # valid internal target "#Acts!A<row>"; pdf mode: the cell is a URL.
        target = index if SOURCE_LINK == "pdf" else f'"#Acts!A"&{index}'
        return f'=IF(ISNA({match}),"",HYPERLINK({target},"{_SOURCE_LABEL}"))'
    return f'=IF(ISNA({match}),"",{index})'


class ExcelWriter:

    def __init__(self, workbook, loader):
        self.wb = workbook
        self.loader = loader

    def _col(self, name):
        """Compliance Database column letter for a field, by name — so the
        Engine formulas never depend on a hardcoded column position."""
        return get_column_letter(
            list(self.loader.df.columns).index(name) + 1
        )

    def _result_row_height(self):
        """Uniform result-row height (pts).  Every result row is the same
        height (a dynamic query can't auto-fit per row), so we size to the
        90th-percentile requirements list — one line per requirement — so
        most rows show fully while the rare huge section (read it in full on
        the Compliance Database sheet) clips instead of making every row
        enormous.  Override with EXCEL_ROW_HEIGHT."""
        override = os.getenv("EXCEL_ROW_HEIGHT")
        if override:
            return min(int(override), _EXCEL_MAX_ROW_HEIGHT)
        chars = 74  # approx chars per line at each Requirements column width
        df = self.loader.df
        cols = [c for c in ("Requirements", "Requirements (2)")
                if c in df.columns]

        def _lines(text):
            if not text:
                return 1
            return sum(max(1, math.ceil(len(line) / chars))
                       for line in str(text).split("\n"))

        line_counts = [1]
        if cols:
            for _, r in df[cols].iterrows():
                # both columns sit on the same row, so size to the taller
                line_counts.append(max(_lines(r[c]) for c in cols))
        line_counts.sort()
        lines = line_counts[int(0.9 * (len(line_counts) - 1))]
        return min(15 * lines + 8, 260)

    def build(self):
        self.write_lists()
        self.write_database()
        self.write_acts()
        self.write_regulators()
        self.write_topics()
        self.write_engine()
        self.write_query()
        self.write_transfer_engine()
        self.write_transfer()
        self.write_home()

    # =================================================================
    # LISTS (hidden) — dropdown sources
    # =================================================================

    def write_lists(self):
        ws = self.wb["Lists"]

        # (header, values) per column; "All" is prepended for filters.
        self._list_ranges = {}
        columns = [
            ("Country", self.loader.country_list),
            ("Topic", self.loader.topic_list),
            ("Data Type", self.loader.datatype_list),
            ("Financial Relevance", self.loader.relevance_list),
            ("Authority", self.loader.regulator_list),
        ]

        for col_idx, (header, values) in enumerate(columns, start=1):
            letter = get_column_letter(col_idx)
            ws.cell(1, col_idx, header)
            entries = ["All"] + list(values)
            for row_idx, value in enumerate(entries, start=2):
                ws.cell(row_idx, col_idx, value)
            last = len(entries) + 1
            self._list_ranges[header] = (
                f"Lists!${letter}$2:${letter}${last}"
            )

    # =================================================================
    # COMPLIANCE DATABASE
    # =================================================================

    def write_database(self):
        ws = self.wb["Compliance Database"]

        columns = list(self.loader.df.columns)
        ws.append(columns)
        for row in self.loader.df.itertuples(index=False):
            ws.append(list(row))

        finish_sheet(ws, "ComplianceTable")

        # Wrap the long text columns and widen them.
        for idx, name in enumerate(columns, start=1):
            if name in WRAP_COLUMNS:
                letter = get_column_letter(idx)
                ws.column_dimensions[letter].width = _col_width(name)
                for cell in ws[letter][1:]:
                    cell.alignment = WRAP

        # The Query INDEXes this column, so keep the raw link target as the
        # value.  Only attach a real hyperlink for external (http) targets —
        # an internal "#'Sheet'!A1" target written this way is invalid and
        # makes Excel strip links on open, which broke "Go to Act".  The
        # Query builds a valid internal HYPERLINK() for those instead.
        if "Source" in columns:
            letter = get_column_letter(columns.index("Source") + 1)
            for cell in ws[letter][1:]:
                v = cell.value
                if isinstance(v, str) and v.startswith("http"):
                    cell.hyperlink = v
                    cell.font = Font(color="0563C1", underline="single")
                elif isinstance(v, (int, float)) and v:
                    # act-row number -> valid internal link (location, not
                    # target, so Excel keeps it).
                    cell.hyperlink = Hyperlink(
                        ref=cell.coordinate, location=f"Acts!A{int(v)}",
                        display="Go to Act")
                    cell.font = Font(color="0563C1", underline="single")
            ws.column_dimensions[letter].width = 12

    # =================================================================
    # QUERY ENGINE
    # =================================================================

    def write_query(self):
        ws = self.wb["Query"]
        ws.sheet_view.showGridLines = False

        ws["A1"] = "DTIA Query Engine"
        ws["A1"].font = TITLE_FONT
        ws["A1"].fill = TITLE_FILL
        ws.merge_cells("A1:D1")

        ws["A3"] = ("Choose filters (leave as 'All' to ignore), then read the "
                    "results below.  Reset: clear cells B4:B10.")
        ws["A3"].font = Font(italic=True)

        # Filter label -> (input cell, Lists header or None for free text)
        filters = [
            ("Country", "B4", "Country"),
            ("Topic", "B5", "Topic"),
            ("Data Type", "B6", "Data Type"),
            ("Financial Relevance", "B7", "Financial Relevance"),
            ("Authority", "B8", "Authority"),
            ("Keyword (free text)", "B9", None),
        ]

        for label, cell, list_header in filters:
            row = int(cell[1:])
            ws.cell(row, 1, label).font = Font(bold=True)
            target = ws[cell]
            if list_header:
                target.value = "All"
                dv = DataValidation(
                    type="list",
                    formula1=self._list_ranges[list_header],
                    allow_blank=True,
                )
                ws.add_data_validation(dv)
                dv.add(target)
            else:
                target.value = ""
            target.fill = HEADER_FILL

        # Yes/No toggle to hide sections with "None" financial relevance.
        ws.cell(10, 1, "Hide non-financial (None)").font = Font(bold=True)
        ws["B10"] = "No"
        dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(ws["B10"])
        ws["B10"].fill = HEADER_FILL

        # Reset "button".  A real one-click button needs a macro (.xlsm),
        # which locked-down Excel installs often block; instead the engine
        # treats a blank filter as "All", so clearing B4:B10 shows
        # everything.  This labelled box makes that gesture discoverable.
        ws.merge_cells("D4:E5")
        reset = ws["D4"]
        reset.value = "↺  Reset filters"
        reset.font = Font(bold=True, color="FFFFFF")
        reset.fill = TITLE_FILL
        reset.alignment = Alignment(horizontal="center", vertical="center")
        for coord in ("D4", "E4", "D5", "E5"):
            ws[coord].border = THIN_BORDER
        ws.merge_cells("D6:E8")
        hint = ws["D6"]
        hint.value = "Select cells B4:B10\nand press Delete"
        hint.font = Font(italic=True, size=9)
        hint.alignment = Alignment(
            horizontal="center", vertical="top", wrap_text=True)

        # Results: classic INDEX/MATCH pulled from the hidden Engine sheet
        # (works in every Excel version — no dynamic arrays).
        n = len(self.loader.df)
        last = n + 1  # data rows on Compliance Database / Engine: 2..n+1

        ws["A12"] = f"Matching sections (of {n} total):"
        ws["A12"].font = Font(bold=True)
        ws["C12"] = f"=Engine!$D$1"          # live match count
        ws["C12"].font = Font(bold=True)

        headers = list(self.loader.df.columns)
        for idx, name in enumerate(headers, start=1):
            c = ws.cell(13, idx, name)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.border = THIN_BORDER

        row_height = self._result_row_height()
        for r in range(14, 14 + n):
            k = r - 13  # this row shows the k-th matching section
            match = f"MATCH({k},Engine!$C$2:$C${last},0)"
            for idx in range(1, len(headers) + 1):
                col = get_column_letter(idx)
                cell = ws.cell(r, idx)
                cell.value = _result_formula(
                    headers[idx - 1], match, col, last
                )
                cell.border = THIN_BORDER
                if headers[idx - 1] in WRAP_COLUMNS:
                    cell.alignment = WRAP
            # Formula cells do not auto-fit height, so size each result row
            # to the busiest requirements list (capped at Excel's limit).
            ws.row_dimensions[r].height = row_height

        for idx, name in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = _col_width(name)
        ws.column_dimensions["A"].width = 16
        ws.freeze_panes = "A14"

    # =================================================================
    # ENGINE (hidden) — per-row match flag + running rank
    # =================================================================

    def write_engine(self):
        ws = self.wb["Engine"]
        n = len(self.loader.df)
        last = n + 1
        db = "'Compliance Database'"

        ws["A1"] = "match"
        ws["B1"] = "rank"
        ws["C1"] = "rankval"
        ws["D1"] = f'=SUM($A$2:$A${last})'   # total matches (shown on Query)

        C = self._col
        for i in range(2, last + 1):
            keyword = '&" "&'.join(
                f"{db}!${C(name)}{i}" for name in (
                    "Heading", "Summary", "DTIA Summary",
                    "Requirements", "Requirements (2)", "Source Quote",
                )
            )
            # Filters: B4 Country, B5 Topic, B6 Data Type,
            # B7 Financial Relevance, B8 Authority, B9 keyword.
            # Each exact-match filter passes when the cell is "All" OR blank
            # (blank = ignored), so clearing B4:B10 resets to "show everything".
            ws.cell(i, 1).value = (
                "=IF(AND("
                f'OR(Query!$B$4="All",Query!$B$4="",'
                f'{db}!${C("Country")}{i}=Query!$B$4),'
                f'OR(Query!$B$5="All",ISNUMBER(SEARCH(Query!$B$5,'
                f'{db}!${C("Topics")}{i}))),'
                f'OR(Query!$B$6="All",ISNUMBER(SEARCH(Query!$B$6,'
                f'{db}!${C("Data Types")}{i}))),'
                f'OR(Query!$B$7="All",Query!$B$7="",'
                f'{db}!${C("Financial Relevance")}{i}=Query!$B$7),'
                f'OR(Query!$B$8="All",Query!$B$8="",'
                f'{db}!${C("Authority")}{i}=Query!$B$8),'
                f'OR(Query!$B$9="",ISNUMBER(SEARCH(Query!$B$9,{keyword}))),'
                f'OR(Query!$B$10<>"Yes",'
                f'{db}!${C("Financial Relevance")}{i}<>"None")'
                "),1,0)"
            )
            # Running rank of matches; C mirrors it as a number for MATCH.
            ws.cell(i, 2).value = f'=IF($A{i}=1,SUM($A$2:$A{i}),"")'
            ws.cell(i, 3).value = f'=IF($A{i}=1,SUM($A$2:$A{i}),0)'

    # =================================================================
    # DATA TRANSFER  —  cross-border view over two countries
    # =================================================================

    def write_transfer(self):
        """Pick two countries; list every section from either one.

        A transfer between A and B is governed by both A's (outbound) and
        B's (inbound) rules, so the result is the union of the two
        countries' sections — grouped by country because the underlying
        data is sorted Country → Act → Section.
        """
        ws = self.wb["Data Transfer"]
        ws.sheet_view.showGridLines = False

        ws["A1"] = "Cross-Border Data Transfer"
        ws["A1"].font = TITLE_FONT
        ws["A1"].fill = TITLE_FILL
        ws.merge_cells("A1:D1")

        ws["A3"] = ("Pick the two countries involved in the transfer. "
                    "Every relevant act/section from either country is "
                    "listed below; add a keyword to narrow it.")
        ws["A3"].font = Font(italic=True)

        # Country dropdowns use the Lists Country column minus the leading
        # "All" (rows 3..last), so both pickers hold real countries.
        n_countries = len(self.loader.country_list)
        country_range = f"Lists!$A$3:$A${n_countries + 2}"

        pickers = [
            ("Country A (from)", "B4", country_range),
            ("Country B (to)", "B5", country_range),
            ("Keyword (free text)", "B6", None),
        ]
        for label, cell, rng in pickers:
            row = int(cell[1:])
            ws.cell(row, 1, label).font = Font(bold=True)
            target = ws[cell]
            if rng:
                dv = DataValidation(type="list", formula1=rng,
                                    allow_blank=True)
                ws.add_data_validation(dv)
                dv.add(target)
            else:
                target.value = ""
            target.fill = HEADER_FILL

        # Default to the first two countries so the sheet is populated on
        # open (leaving Country B blank simply shows Country A alone).
        countries = self.loader.country_list
        ws["B4"] = countries[0] if countries else ""
        ws["B5"] = countries[1] if len(countries) > 1 else ""

        n = len(self.loader.df)
        last = n + 1

        ws["A8"] = f"Matching sections (of {n} total):"
        ws["A8"].font = Font(bold=True)
        ws["C8"] = "='Transfer Engine'!$D$1"
        ws["C8"].font = Font(bold=True)

        headers = list(self.loader.df.columns)
        for idx, name in enumerate(headers, start=1):
            c = ws.cell(9, idx, name)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.border = THIN_BORDER

        row_height = self._result_row_height()
        for r in range(10, 10 + n):
            k = r - 9  # k-th matching section
            match = f"MATCH({k},'Transfer Engine'!$C$2:$C${last},0)"
            for idx in range(1, len(headers) + 1):
                col = get_column_letter(idx)
                cell = ws.cell(r, idx)
                cell.value = _result_formula(
                    headers[idx - 1], match, col, last
                )
                cell.border = THIN_BORDER
                if headers[idx - 1] in WRAP_COLUMNS:
                    cell.alignment = WRAP
            ws.row_dimensions[r].height = row_height

        for idx, name in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = _col_width(name)
        ws.column_dimensions["A"].width = 18
        ws.freeze_panes = "A10"

    # =================================================================
    # TRANSFER ENGINE (hidden) — union-of-two-countries match flag
    # =================================================================

    def write_transfer_engine(self):
        ws = self.wb["Transfer Engine"]
        n = len(self.loader.df)
        last = n + 1
        db = "'Compliance Database'"
        q = "'Data Transfer'"

        ws["A1"] = "match"
        ws["B1"] = "rank"
        ws["C1"] = "rankval"
        ws["D1"] = f'=SUM($A$2:$A${last})'

        C = self._col
        country = C("Country")
        for i in range(2, last + 1):
            keyword = '&" "&'.join(
                f"{db}!${C(name)}{i}" for name in (
                    "Heading", "Summary", "DTIA Summary",
                    "Requirements", "Requirements (2)", "Source Quote",
                )
            )
            # Row kept if its Country is either selected country and the
            # optional keyword hits its text.
            ws.cell(i, 1).value = (
                "=IF(AND("
                f'OR({db}!${country}{i}={q}!$B$4,'
                f'{db}!${country}{i}={q}!$B$5),'
                f'OR({q}!$B$6="",ISNUMBER(SEARCH({q}!$B$6,{keyword})))'
                "),1,0)"
            )
            ws.cell(i, 2).value = f'=IF($A{i}=1,SUM($A$2:$A{i}),"")'
            ws.cell(i, 3).value = f'=IF($A{i}=1,SUM($A$2:$A{i}),0)'

    # =================================================================
    # HOME
    # =================================================================

    def write_home(self):
        ws = self.wb["Home"]
        ws.sheet_view.showGridLines = False

        ws["A1"] = "DTIA Compliance Knowledge Base"
        ws["A1"].font = TITLE_FONT
        ws["A1"].fill = TITLE_FILL
        ws.merge_cells("A1:C1")

        stats = self.loader.summary()
        rows = [
            ("Countries", stats["countries"]),
            ("Acts", stats["acts"]),
            ("Regulators", stats["regulators"]),
            ("Topics", stats["topics"]),
            ("Compliance sections", stats["rows"]),
        ]
        for i, (label, value) in enumerate(rows, start=3):
            ws.cell(i, 1, label).font = Font(bold=True)
            ws.cell(i, 2, value)

        ws["A10"] = "How to use"
        ws["A10"].font = Font(bold=True, size=13)
        steps = [
            "1. Open the Query sheet.",
            "2. Pick a Country, Category, Topic, Data Type, "
            "Financial Relevance or Authority — or type a keyword.",
            "3. Matching sections appear below, each with its Act and "
            "page reference for the source document.",
            "4. The Data Transfer sheet answers cross-border questions: "
            "pick two countries and see every relevant act/section from "
            "either side of the transfer.",
            "5. The Compliance Database sheet holds every extracted "
            "section; Acts / Regulators / Topics summarise them.",
        ]
        for i, step in enumerate(steps, start=11):
            ws.cell(i, 1, step).alignment = Alignment(wrap_text=True)
            ws.merge_cells(f"A{i}:C{i}")

        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 14

    # =================================================================
    # ACTS
    # =================================================================

    def write_acts(self):
        ws = self.wb["Acts"]
        ws.append([
            "Country", "Act", "Sections", "Regulators", "Topics",
            "Financial Relevance", "Source PDF",
        ])
        for act in self.loader.acts.values():
            ws.append([
                act["Country"], act["Act"], act["Sections"],
                ", ".join(sorted(act["Regulators"])),
                ", ".join(sorted(act["Topics"])),
                act["Financial Relevance"],
                _pdf_url(act["Country"], act["Act"]),
            ])
        finish_sheet(ws, "ActsTable")

        # The Query's "Go to Act" link lands here; each act's PDF link lives
        # in this column (external URL → clickable "Open PDF").
        pdf_col = get_column_letter(7)
        for cell in ws[pdf_col][1:]:
            if cell.value:
                cell.hyperlink = cell.value
                cell.value = "Open PDF"
                cell.font = Font(color="0563C1", underline="single")
        ws.column_dimensions[pdf_col].width = 12

    # =================================================================
    # REGULATORS
    # =================================================================

    def write_regulators(self):
        ws = self.wb["Regulators"]
        ws.append(["Authority", "Countries", "Acts"])
        for regulator, info in sorted(self.loader.regulators.items()):
            ws.append([
                regulator,
                ", ".join(sorted(info["Country"])),
                ", ".join(sorted(info["Acts"])),
            ])
        finish_sheet(ws, "RegulatorTable")

    # =================================================================
    # TOPICS
    # =================================================================

    def write_topics(self):
        ws = self.wb["Topics"]
        ws.append(["Topic", "Sections"])
        for topic, count in sorted(
            self.loader.topics.items(), key=lambda kv: -kv[1]
        ):
            ws.append([topic, count])
        finish_sheet(ws, "TopicTable")
