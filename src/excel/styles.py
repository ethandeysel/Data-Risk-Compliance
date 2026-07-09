from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment
)

# --------------------------------------------------
# Colours
# --------------------------------------------------

NAVY = "1F4E78"
GREEN = "D9EAD3"
GREY = "EEEEEE"
WHITE = "FFFFFF"

# Confidence colours
HIGH = "C6EFCE"
MEDIUM = "FFEB9C"
LOW = "F4CCCC"

# --------------------------------------------------
# Fonts
# --------------------------------------------------

TITLE_FONT = Font(
    bold=True,
    size=16,
    color=WHITE
)

HEADER_FONT = Font(
    bold=True,
    size=11
)

BODY_FONT = Font(
    size=10
)

# --------------------------------------------------
# Fills
# --------------------------------------------------

TITLE_FILL = PatternFill(
    fill_type="solid",
    fgColor=NAVY
)

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor=GREEN
)

GREY_FILL = PatternFill(
    fill_type="solid",
    fgColor=GREY
)

HIGH_FILL = PatternFill(
    fill_type="solid",
    fgColor=HIGH
)

MEDIUM_FILL = PatternFill(
    fill_type="solid",
    fgColor=MEDIUM
)

LOW_FILL = PatternFill(
    fill_type="solid",
    fgColor=LOW
)

# --------------------------------------------------
# Borders
# --------------------------------------------------

THIN_BORDER = Border(

    left=Side(style="thin"),

    right=Side(style="thin"),

    top=Side(style="thin"),

    bottom=Side(style="thin")

)

# --------------------------------------------------
# Alignment
# --------------------------------------------------

CENTER = Alignment(
    horizontal="center",
    vertical="center",
    wrap_text=True
)

WRAP = Alignment(
    vertical="top",
    wrap_text=True
)

LEFT = Alignment(
    horizontal="left",
    vertical="top",
    wrap_text=True
)